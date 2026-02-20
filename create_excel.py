#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "Executive Summary"

title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
header_font = Font(name='Calibri', size=12, bold=True)
normal_font = Font(name='Calibri', size=10)
title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_col_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def add_title(ws, text, row):
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = title_font
    ws.cell(row=row, column=1).fill = title_fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

def add_header(ws, text, row):
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

def add_cell(ws, text, row, col):
    ws.cell(row=row, column=col, value=text)
    ws.cell(row=row, column=col).font = normal_font
    ws.cell(row=row, column=col).border = thin_border
    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

row = 1
add_title(ws, "Patch Management Strategy - Executive Summary", row)
row += 1

ws.cell(row=row, column=1, value="Document Version:")
ws.cell(row=row, column=2, value="1.0")
row += 1
ws.cell(row=row, column=1, value="Effective Date:")
ws.cell(row=row, column=2, value="February 2026")
row += 1
ws.cell(row=row, column=1, value="Document Owner:")
ws.cell(row=row, column=2, value="Enterprise Architecture / IT Operations")
row += 2

add_header(ws, "Strategy Overview", row)
row += 1

summary_text = """This document establishes the enterprise patch management strategy for a three-environment system (DEV, TEST, PROD). It defines the governance framework, operational procedures, risk mitigation controls, and decision criteria required to maintain system security, stability, and compliance while enabling efficient delivery of patches across environments."""

ws.cell(row=row, column=1, value=summary_text)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

add_header(ws, "Key Assumptions", row)
row += 1

assumptions = [
    "DEV, TEST, and PROD environments are logically and physically separated",
    "Change Management Board (CAB) exists with authority to approve/prohibit changes",
    "Production changes follow a formal Change Advisory Board (CAB) process",
    "Automated deployment tooling is available (e.g., CI/CD pipelines)",
    "Backup and recovery procedures are established for all environments",
    "The organization has defined SLAs for system availability"
]

for a in assumptions:
    ws.cell(row=row, column=1, value="• " + a)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

row += 1
add_header(ws, "Patch Deployment Sequence", row)
row += 1

headers = ["Sequence", "Environment", "Purpose", "Typical Lead Time"]
for col, h in enumerate(headers, 1):
    ws.cell(row=row, column=col, value=h)
    ws.cell(row=row, column=col).font = header_font
    ws.cell(row=row, column=col).fill = header_fill
    ws.cell(row=row, column=col).border = thin_border
row += 1

deploy_seq = [
    ("1", "DEV", "Initial validation, development testing", "Day 0"),
    ("2", "TEST / STAGING", "Full regression, UAT", "Day 1-3"),
    ("3", "PROD", "Production release", "Day 5-7")
]
for seq in deploy_seq:
    for col, val in enumerate(seq, 1):
        ws.cell(row=row, column=col, value=val)
        ws.cell(row=row, column=col).border = thin_border
    row += 1

set_col_widths(ws, [15, 20, 35, 20])

# Sheet 2: Patch Promotion Flow
ws2 = wb.create_sheet("Patch Promotion Flow")
row = 1

add_title(ws2, "Patch Promotion Flow & Criteria", row)
row += 2

add_header(ws2, "Entry Criteria per Environment", row)
row += 1

entry_data = [
    ("DEV", "Patch package validated by development team; change request created; backup verified"),
    ("TEST", "DEV deployment successful; test cases updated; test data prepared; test environment availability confirmed"),
    ("PROD", "TEST execution results approved; CAB approval obtained; rollback plan documented; communication plan initiated")
]

ws2.cell(row=row, column=1, value="Environment")
ws2.cell(row=row, column=2, value="Entry Criteria")
for col in range(1, 3):
    ws2.cell(row=row, column=col).font = header_font
    ws2.cell(row=row, column=col).fill = header_fill
    ws2.cell(row=row, column=col).border = thin_border
row += 1

for env, criteria in entry_data:
    ws2.cell(row=row, column=1, value=env)
    ws2.cell(row=row, column=2, value=criteria)
    ws2.cell(row=row, column=1).border = thin_border
    ws2.cell(row=row, column=1).font = header_font
    ws2.cell(row=row, column=2).border = thin_border
    ws2.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    row += 1

row += 1
add_header(ws2, "Exit Criteria per Environment", row)
row += 1

exit_data = [
    ("DEV", "Unit tests pass (≥90% coverage); basic integration smoke tests pass; no blocking defects"),
    ("TEST", "All blocking/critical defects resolved; regression test suite pass rate ≥95%; UAT sign-off obtained; performance benchmarks met"),
    ("PROD", "Post-deployment smoke tests pass; monitoring alerts confirmed; business stakeholders notified; documentation updated")
]

ws2.cell(row=row, column=1, value="Environment")
ws2.cell(row=row, column=2, value="Exit Criteria")
for col in range(1, 3):
    ws2.cell(row=row, column=col).font = header_font
    ws2.cell(row=row, column=col).fill = header_fill
    ws2.cell(row=row, column=col).border = thin_border
row += 1

for env, criteria in exit_data:
    ws2.cell(row=row, column=1, value=env)
    ws2.cell(row=row, column=2, value=criteria)
    ws2.cell(row=row, column=1).border = thin_border
    ws2.cell(row=row, column=1).font = header_font
    ws2.cell(row=row, column=2).border = thin_border
    ws2.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    row += 1

row += 1
add_header(ws2, "Rollback Strategy per Environment", row)
row += 1

rollback_headers = ["Environment", "Rollback Trigger", "Rollback Procedure", "Max Downtime Tolerance"]
for col, h in enumerate(rollback_headers, 1):
    ws2.cell(row=row, column=col, value=h)
    ws2.cell(row=row, column=col).font = header_font
    ws2.cell(row=row, column=col).fill = header_fill
    ws2.cell(row=row, column=col).border = thin_border
row += 1

rollback_data = [
    ("DEV", "Any deployment failure", "Redeploy previous baseline from version control", "1 hour"),
    ("TEST", "Critical defects discovered", "Restore from TEST backup; redeploy previous version", "4 hours"),
    ("PROD", "Service availability <99.5%; data integrity issues; critical defect affecting >=5% of users", "Execute documented backout plan; engage on-call DBA if needed", "Per SLA (15-30 min)")
]

for r in rollback_data:
    for col, val in enumerate(r, 1):
        ws2.cell(row=row, column=col, value=val)
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
    row += 1

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 25

# Sheet 3: Testing Requirements
ws3 = wb.create_sheet("Testing Requirements")
row = 1

add_title(ws3, "Testing & Validation Requirements", row)
row += 2

add_header(ws3, "Regression Testing Requirements", row)
row += 1

reg_headers = ["Scope", "Coverage", "Ownership", "Tooling"]
for col, h in enumerate(reg_headers, 1):
    ws3.cell(row=row, column=col, value=h)
    ws3.cell(row=row, column=col).font = header_font
    ws3.cell(row=row, column=col).fill = header_fill
    ws3.cell(row=row, column=col).border = thin_border
row += 1

reg_data = [
    ("Full regression", "100% of functional test cases", "QA Team", "Automated test framework"),
    ("Targeted regression", "Affected components + downstream dependencies", "QA Team + Development", "Automated + manual"),
    ("Ad-hoc / exploratory", "Boundary conditions, edge cases", "QA Team", "Manual")
]

for r in reg_data:
    for col, val in enumerate(r, 1):
        ws3.cell(row=row, column=col, value=val)
        ws3.cell(row=row, column=col).border = thin_border
        ws3.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
    row += 1

row += 1
add_header(ws3, "Smoke Testing vs Full Regression Criteria", row)
row += 1

smoke_headers = ["Factor", "Smoke Test", "Full Regression"]
for col, h in enumerate(smoke_headers, 1):
    ws3.cell(row=row, column=col, value=h)
    ws3.cell(row=row, column=col).font = header_font
    ws3.cell(row=row, column=col).fill = header_fill
    ws3.cell(row=row, column=col).border = thin_border
row += 1

smoke_data = [
    ("When Used", "Every deployment", "Weekly or per release"),
    ("Scope", "Critical path (5-10 tests)", "Complete suite (200+ tests)"),
    ("Execution Time", "<30 minutes", "4-8 hours"),
    ("Ownership", "Operations + Dev", "QA Team"),
    ("Automation", "Fully automated", "Automated + manual"),
    ("Pass Threshold", "100% pass required", "≥95% pass required")
]

for r in smoke_data:
    for col, val in enumerate(r, 1):
        ws3.cell(row=row, column=col, value=val)
        ws3.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
add_header(ws3, "UAT Requirements", row)
row += 1

uat_data = [
    ("UAT Sign-off", "Required for ALL production deployments"),
    ("UAT Participants", "Business process owners, key end users"),
    ("UAT Test Cases", "Business-critical scenarios only (minimum 15 scenarios)"),
    ("UAT Duration", "Minimum 1 business day for standard patches; 3-5 business days for major releases"),
    ("UAT Defect Severity", "Blocking/critical defects must be resolved before PROD promotion")
]

for criterion, req in uat_data:
    ws3.cell(row=row, column=1, value=criterion)
    ws3.cell(row=row, column=2, value=req)
    ws3.cell(row=row, column=1).border = thin_border
    ws3.cell(row=row, column=1).font = header_font
    ws3.cell(row=row, column=2).border = thin_border
    ws3.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    row += 1

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 45
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 25

# Sheet 4: Patch Classification
ws4 = wb.create_sheet("Patch Classification")
row = 1

add_title(ws4, "Patch Categorization Framework", row)
row += 2

add_header(ws4, "Patch Classification Definitions", row)
row += 1

class_headers = ["Classification", "Definition", "Response SLA", "Deployment Target"]
for col, h in enumerate(class_headers, 1):
    ws4.cell(row=row, column=col, value=h)
    ws4.cell(row=row, column=col).font = header_font
    ws4.cell(row=row, column=col).fill = header_fill
    ws4.cell(row=row, column=col).border = thin_border
row += 1

class_data = [
    ("P1-Critical", "Patches addressing active exploits, critical CVEs (CVSS ≥7.0), or confirmed security incidents", "24-48 hours (critical: 4-24 hours)", "4-24 hours"),
    ("P2-High", "Significant security risk, major functionality impact", "4 hours response, 7 days resolution", "5-7 days"),
    ("P3-Medium", "Moderate impact, workaround available", "24 hours response", "Monthly cadence"),
    ("P4-Low", "Minimal impact, cosmetic issues, documentation updates", "5 business days", "Quarterly")
]

for r in class_data:
    for col, val in enumerate(r, 1):
        ws4.cell(row=row, column=col, value=val)
        ws4.cell(row=row, column=col).border = thin_border
        ws4.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
    row += 1

row += 1
add_header(ws4, "Patch Type Testing Scope", row)
row += 1

type_headers = ["Category", "Testing Scope", "Approval Level", "Rollback Complexity"]
for col, h in enumerate(type_headers, 1):
    ws4.cell(row=row, column=col, value=h)
    ws4.cell(row=row, column=col).font = header_font
    ws4.cell(row=row, column=col).fill = header_fill
    ws4.cell(row=row, column=col).border = thin_border
row += 1

type_data = [
    ("Infrastructure", "Reduced - focus on availability", "Standard CAB", "High"),
    ("Database", "Moderate - data integrity focus", "DBA Lead + CAB", "Critical"),
    ("Middleware", "Moderate - integration focus", "Standard CAB", "Medium"),
    ("Application", "Full - functional + integration", "Standard CAB", "Low")
]

for r in type_data:
    for col, val in enumerate(r, 1):
        ws4.cell(row=row, column=col, value=val)
        ws4.cell(row=row, column=col).border = thin_border
        ws4.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
    row += 1

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 35
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 20

# Sheet 5: Governance Controls
ws5 = wb.create_sheet("Governance Controls")
row = 1

add_title(ws5, "Governance & Controls", row)
row += 2

add_header(ws5, "Change Management Approvals per Environment", row)
row += 1

approval_headers = ["Environment", "Approval Required", "Approver", "Lead Time"]
for col, h in enumerate(approval_headers, 1):
    ws5.cell(row=row, column=col, value=h)
    ws5.cell(row=row, column=col).font = header_font
    ws5.cell(row=row, column=col).fill = header_fill
    ws5.cell(row=row, column=col).border = thin_border
row += 1

approval_data = [
    ("DEV", "Notification only", "Development Lead", "Same day"),
    ("TEST", "Change request approval", "Test Manager", "24 hours"),
    ("PROD", "CAB approval (standard)", "CAB (majority)", "5 business days"),
    ("PROD", "Emergency approval", "CAB Chair + Security Lead", "4-24 hours")
]

for r in approval_data:
    for col, val in enumerate(r, 1):
        ws5.cell(row=row, column=col, value=val)
        ws5.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
add_header(ws5, "Communication Plan", row)
row += 1

comm_headers = ["Stakeholder", "Notification Trigger", "Timing", "Channel"]
for col, h in enumerate(comm_headers, 1):
    ws5.cell(row=row, column=col, value=h)
    ws5.cell(row=row, column=col).font = header_font
    ws5.cell(row=row, column=col).fill = header_fill
    ws5.cell(row=row, column=col).border = thin_border
row += 1

comm_data = [
    ("Development Team", "Patch available in DEV", "Upon release", "Teams / Email"),
    ("QA Team", "Deployment to TEST", "24 hrs before", "Teams / Email"),
    ("Business Stakeholders", "PROD deployment planned", "5 business days before", "Email"),
    ("End Users", "PROD deployment", "48 hours before (planned)", "Portal / Email"),
    ("CAB", "All changes", "Per approval timeline", "Agenda + Email"),
    ("IT Operations", "All deployments", "24 hours before", "Teams / PagerDuty"),
    ("Security Team", "Security patches", "Immediate", "Teams / Phone")
]

for r in comm_data:
    for col, val in enumerate(r, 1):
        ws5.cell(row=row, column=col, value=val)
        ws5.cell(row=row, column=col).border = thin_border
        ws5.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
    row += 1

row += 1
add_header(ws5, "Documentation Requirements", row)
row += 1

doc_headers = ["Document", "Required For", "Retention"]
for col, h in enumerate(doc_headers, 1):
    ws5.cell(row=row, column=col, value=h)
    ws5.cell(row=row, column=col).font = header_font
    ws5.cell(row=row, column=col).fill = header_fill
    ws5.cell(row=row, column=col).border = thin_border
row += 1

doc_data = [
    ("Change Request (CR)", "All environments", "7 years"),
    ("Test Results Report", "TEST + PROD", "5 years"),
    ("UAT Sign-off", "PROD", "5 years"),
    ("Rollback Procedure", "PROD", "5 years"),
    ("Post-Implementation Review (PIR)", "PROD", "7 years"),
    ("Security Impact Assessment", "Security patches, major upgrades", "7 years")
]

for r in doc_data:
    for col, val in enumerate(r, 1):
        ws5.cell(row=row, column=col, value=val)
        ws5.cell(row=row, column=col).border = thin_border
    row += 1

ws5.column_dimensions['A'].width = 22
ws5.column_dimensions['B'].width = 35
ws5.column_dimensions['C'].width = 22
ws5.column_dimensions['D'].width = 22

# Sheet 6: RACI Matrix
ws6 = wb.create_sheet("RACI Matrix")
row = 1

add_title(ws6, "RACI Matrix for Patch Activities", row)
row += 2

raci_headers = ["Activity", "IT Ops", "Development", "QA", "Security", "CAB", "Business"]
for col, h in enumerate(raci_headers, 1):
    ws6.cell(row=row, column=col, value=h)
    ws6.cell(row=row, column=col).font = header_font
    ws6.cell(row=row, column=col).fill = header_fill
    ws6.cell(row=row, column=col).border = thin_border
row += 1

raci_data = [
    ("Patch source identification", "R", "C", "I", "R", "I", "-"),
    ("Patch testing in DEV", "C", "R", "C", "I", "-", "-"),
    ("Change request creation", "R", "C", "C", "C", "I", "I"),
    ("TEST environment deployment", "R", "C", "C", "I", "-", "-"),
    ("Regression test execution", "C", "C", "R", "I", "-", "-"),
    ("UAT coordination", "I", "C", "R", "I", "I", "R"),
    ("CAB approval", "I", "C", "C", "C", "R", "I"),
    ("PROD deployment execution", "R", "C", "I", "C", "I", "-"),
    ("Post-deployment validation", "R", "C", "C", "C", "-", "-"),
    ("Post-implementation review", "R", "C", "C", "C", "C", "I"),
    ("Rollback execution", "R", "C", "-", "C", "I", "-")
]

for r in raci_data:
    for col, val in enumerate(r, 1):
        ws6.cell(row=row, column=col, value=val)
        ws6.cell(row=row, column=col).border = thin_border
        ws6.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1

row += 1
ws6.cell(row=row, column=1, value="Legend: R=Responsible, A=Accountable, C=Consulted, I=Informed")
ws6.cell(row=row, column=1).font = Font(name='Calibri', size=9, italic=True)
ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)

ws6.column_dimensions['A'].width = 30
for col in range(2, 8):
    ws6.column_dimensions[get_column_letter(col)].width = 12

# Sheet 7: Templates
ws7 = wb.create_sheet("Templates & Checklists")
row = 1

add_title(ws7, "Templates & Checklists", row)
row += 2

add_header(ws7, "Entry/Exit Criteria Checklist", row)
row += 1

checklist_headers = ["Phase", "Checklist Item", "Status"]
for col, h in enumerate(checklist_headers, 1):
    ws7.cell(row=row, column=col, value=h)
    ws7.cell(row=row, column=col).font = header_font
    ws7.cell(row=row, column=col).fill = header_fill
    ws7.cell(row=row, column=col).border = thin_border
row += 1

checklist = [
    ("Pre-DEV", "Change request created and assigned", ""),
    ("Pre-DEV", "Patch package validated by development team", ""),
    ("Pre-DEV", "Backup verified and documented", ""),
    ("Pre-TEST", "DEV deployment successful", ""),
    ("Pre-TEST", "Test cases updated for new functionality", ""),
    ("Pre-TEST", "Regression test suite ready", ""),
    ("Pre-PROD", "TEST execution results approved", ""),
    ("Pre-PROD", "Regression test pass rate ≥95%", ""),
    ("Pre-PROD", "UAT sign-off obtained", ""),
    ("Pre-PROD", "CAB approval obtained", ""),
    ("Pre-PROD", "Rollback plan documented and tested", ""),
    ("Post-PROD", "Post-deployment smoke tests pass", ""),
    ("Post-PROD", "Monitoring alerts confirmed operational", ""),
    ("Post-PROD", "Database integrity confirmed", ""),
    ("Post-PROD", "Business stakeholders notified", "")
]

for phase, item, status in checklist:
    ws7.cell(row=row, column=1, value=phase)
    ws7.cell(row=row, column=2, value=item)
    ws7.cell(row=row, column=3, value=status)
    for col in range(1, 4):
        ws7.cell(row=row, column=col).border = thin_border
    row += 1

ws7.column_dimensions['A'].width = 15
ws7.column_dimensions['B'].width = 50
ws7.column_dimensions['C'].width = 12

# Sheet 8: SOP Overview
ws8 = wb.create_sheet("SOP - Overview")
row = 1

add_title(ws8, "SOP - Playbook Overview", row)
row += 2

ws8.cell(row=row, column=1, value="Purpose and Scope")
ws8.cell(row=row, column=1).font = header_font
row += 1

scope_text = """This operational playbook provides step-by-step procedures for executing patch deployments across the three-environment system (DEV → TEST → PROD). It translates the Patch Management Strategy into actionable tasks for operations teams."""

ws8.cell(row=row, column=1, value=scope_text)
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws8.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
row += 2

add_header(ws8, "Target Audience", row)
row += 1

audience_headers = ["Role", "Use This Playbook For"]
for col, h in enumerate(audience_headers, 1):
    ws8.cell(row=row, column=col, value=h)
    ws8.cell(row=row, column=col).font = header_font
    ws8.cell(row=row, column=col).fill = header_fill
    ws8.cell(row=row, column=col).border = thin_border
row += 1

audience = [
    ("IT Operations Engineers", "Executing deployments, validations, rollbacks"),
    ("DevOps Engineers", "Pipeline execution, automation troubleshooting"),
    ("Help Desk/Support Staff", "Understanding deployment status, user communication"),
    ("Test Engineers", "TEST environment deployment and validation"),
    ("On-Call Engineers", "Emergency response, incident handling")
]

for role, use_for in audience:
    ws8.cell(row=row, column=1, value=role)
    ws8.cell(row=row, column=2, value=use_for)
    ws8.cell(row=row, column=1).border = thin_border
    ws8.cell(row=row, column=2).border = thin_border
    row += 1

row += 1
add_header(ws8, "When to Use This Playbook", row)
row += 1

use_text = """Use this playbook when:
• Deploying routine monthly patches
• Applying security patches (planned or emergency)
• Executing major version upgrades
• Performing emergency patch response
• Executing rollback procedures
• Validating post-deployment system state"""

ws8.cell(row=row, column=1, value=use_text)
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws8.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

ws8.column_dimensions['A'].width = 25
ws8.column_dimensions['B'].width = 50

# Sheet 9: SOP Deployment Steps
ws9 = wb.create_sheet("SOP - Deployment Steps")
row = 1

add_title(ws9, "SOP - Deployment Procedures", row)
row += 2

add_header(ws9, "DEV Deployment Steps", row)
row += 1

dev_steps = [
    "1. Prepare the deployment environment",
    "2. Stop dependent services",
    "3. Execute pre-deployment backup",
    "4. Apply the patch",
    "5. Start services",
    "6. Execute smoke tests",
    "7. Verify basic functionality",
    "8. Document deployment results"
]

for step in dev_steps:
    ws9.cell(row=row, column=1, value=step)
    ws9.cell(row=row, column=1).border = thin_border
    row += 1

row += 1
add_header(ws9, "TEST Deployment Steps", row)
row += 1

test_steps = [
    "1. Verify DEV success before promoting to TEST",
    "2. Prepare the TEST deployment environment",
    "3. Stop TEST environment services",
    "4. Execute pre-deployment backup",
    "5. Apply the patch",
    "6. Start services and verify",
    "7. Execute smoke tests (100% pass required)",
    "8. Execute regression test suite",
    "9. Notify QA team for manual testing",
    "10. Obtain UAT sign-off"
]

for step in test_steps:
    ws9.cell(row=row, column=1, value=step)
    ws9.cell(row=row, column=1).border = thin_border
    row += 1

row += 1
add_header(ws9, "PROD Deployment Steps", row)
row += 1

prod_steps = [
    "1. Final pre-deployment verification",
    "2. Send deployment start notification",
    "3. Enter maintenance window",
    "4. Stop production services",
    "5. Execute production backup",
    "6. Apply the patch",
    "7. Verify patch application",
    "8. Start production services",
    "9. Execute post-deployment smoke tests",
    "10. Verify monitoring and alerting",
    "11. Execute data integrity checks",
    "12. Send deployment completion notification"
]

for step in prod_steps:
    ws9.cell(row=row, column=1, value=step)
    ws9.cell(row=row, column=1).border = thin_border
    row += 1

ws9.column_dimensions['A'].width = 60

# Sheet 10: SOP Rollback
ws10 = wb.create_sheet("SOP - Rollback")
row = 1

add_title(ws10, "SOP - Rollback Procedures", row)
row += 2

add_header(ws10, "Decision Triggers (When to Rollback)", row)
row += 1

trigger_headers = ["Trigger", "Threshold", "Action"]
for col, h in enumerate(trigger_headers, 1):
    ws10.cell(row=row, column=col, value=h)
    ws10.cell(row=row, column=col).font = header_font
    ws10.cell(row=row, column=col).fill = header_fill
    ws10.cell(row=row, column=col).border = thin_border
row += 1

triggers = [
    ("Service Availability", "<99.5%", "Immediate rollback"),
    ("Critical Defects", "≥5% users affected", "Immediate rollback"),
    ("Data Integrity", "Any confirmed issue", "Immediate rollback"),
    ("Security Vulnerability", "CVSS ≥7.0 discovered", "Fast-track rollback"),
    ("Smoke Test Failure", "Any failure in PROD", "Evaluate + decide"),
    ("Performance Degradation", ">20% slower", "Evaluate impact")
]

for t in triggers:
    for col, val in enumerate(t, 1):
        ws10.cell(row=row, column=col, value=val)
        ws10.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
add_header(ws10, "Rollback Decision Authority", row)
row += 1

auth_data = [
    ("DEV", "Lead Developer"),
    ("TEST", "Test Manager + Development Lead (joint)"),
    ("PROD", "CAB Chair + IT Operations Manager (joint)")
]

for env, authority in auth_data:
    ws10.cell(row=row, column=1, value=env)
    ws10.cell(row=row, column=2, value=authority)
    ws10.cell(row=row, column=1).border = thin_border
    ws10.cell(row=row, column=1).font = header_font
    ws10.cell(row=row, column=2).border = thin_border
    row += 1

row += 1
add_header(ws10, "Rollback Procedure Summary", row)
row += 1

prod_rollback = [
    "1. Confirm rollback decision - Obtain required approvals (CAB Chair + IT Ops Manager)",
    "2. Send rollback notification",
    "3. Enter maintenance mode",
    "4. Stop production services",
    "5. Restore production backup",
    "6. Start production services",
    "7. Verify rollback success",
    "8. Send rollback completion notification"
]

for step in prod_rollback:
    ws10.cell(row=row, column=1, value=step)
    ws10.cell(row=row, column=1).border = thin_border
    row += 1

ws10.column_dimensions['A'].width = 25
ws10.column_dimensions['B'].width = 50

# Sheet 11: SOP Troubleshooting
ws11 = wb.create_sheet("SOP - Troubleshooting")
row = 1

add_title(ws11, "SOP - Troubleshooting Guide", row)
row += 2

add_header(ws11, "Patch Deployment Failures", row)
row += 1

ws11.cell(row=row, column=1, value="Issue: Deployment script fails")
ws11.cell(row=row, column=1).font = header_font
row += 1

fix_steps = [
    "1. Check error message in output",
    "2. Review deployment logs: tail -100 /var/log/[deployment].log",
    "3. Verify file permissions: ls -la /opt/[application]",
    "4. Verify disk space: df -h",
    "5. Check dependencies: ldd /opt/[application]/bin/[binary]",
    "6. Retry deployment if transient error"
]

for step in fix_steps:
    ws11.cell(row=row, column=1, value=step)
    ws11.cell(row=row, column=1).border = thin_border
    row += 1

row += 1
ws11.cell(row=row, column=1, value="Issue: Service won't start after patch")
ws11.cell(row=row, column=1).font = header_font
row += 1

fix_steps2 = [
    "1. Check service status: systemctl status [service]",
    "2. Review startup logs: journalctl -u [service] -n 100",
    "3. Verify configuration syntax: nginx -t",
    "4. Check port conflicts: ss -tuln | grep [port]",
    "5. Verify file permissions",
    "6. Check dependencies: ldd on libraries",
    "7. Rollback if unresolvable"
]

for step in fix_steps2:
    ws11.cell(row=row, column=1, value=step)
    ws11.cell(row=row, column=1).border = thin_border
    row += 1

row += 1
add_header(ws11, "Performance Degradation", row)
row += 1

ws11.cell(row=row, column=1, value="Issue: Application slow after patch")
ws11.cell(row=row, column=1).font = header_font
row += 1

perf_steps = [
    "1. Check resource utilization: top, free -h, iostat",
    "2. Compare with baseline metrics",
    "3. Check database query performance",
    "4. Verify index usage",
    "5. Check for new bottlenecks",
    "6. Enable slow query logging if database",
    "7. Rollback if >20% degradation"
]

for step in perf_steps:
    ws11.cell(row=row, column=1, value=step)
    ws11.cell(row=row, column=1).border = thin_border
    row += 1

row += 1
add_header(ws11, "Data Integrity Concerns", row)
row += 1

ws11.cell(row=row, column=1, value="Issue: Data inconsistency after patch")
ws11.cell(row=row, column=1).font = header_font
row += 1

data_steps = [
    "1. HALT deployment immediately",
    "2. Do NOT make any changes",
    "3. Verify data with pre-deployment backup",
    "4. Compare record counts",
    "5. Check referential integrity",
    "6. Execute ROLLBACK immediately",
    "7. Escalate to DBA Lead"
]

for step in data_steps:
    ws11.cell(row=row, column=1, value=step)
    ws11.cell(row=row, column=1).border = thin_border
    row += 1

ws11.column_dimensions['A'].width = 60

# Sheet 12: Emergency Procedures
ws12 = wb.create_sheet("SOP - Emergency")
row = 1

add_title(ws12, "SOP - Emergency Patch Procedures", row)
row += 2

add_header(ws12, "Trigger Criteria for Emergency Classification", row)
row += 1

emergency_headers = ["Criterion", "Threshold", "Example"]
for col, h in enumerate(emergency_headers, 1):
    ws12.cell(row=row, column=col, value=h)
    ws12.cell(row=row, column=col).font = header_font
    ws12.cell(row=row, column=col).fill = header_fill
    ws12.cell(row=row, column=col).border = thin_border
row += 1

emergency = [
    ("Active Exploitation", "Confirmed in-the-wild exploit", "CVE with public PoC"),
    ("CVSS Score", "≥9.0 (Critical)", "Remote code execution"),
    ("Service Availability", "Immediate risk", "Ransomware indicators"),
    ("Regulatory Directive", "Required immediate action", "Compliance enforcement"),
    ("Business Impact", "Complete system outage potential", "Critical system down")
]

for e in emergency:
    for col, val in enumerate(e, 1):
        ws12.cell(row=row, column=col, value=val)
        ws12.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
add_header(ws12, "Accelerated Approval Process", row)
row += 1

accelerated = [
    "1. Contact emergency approvers in parallel: CAB Chair, Security Lead, IT Operations Manager",
    "2. Provide emergency approval request with CVSS score and risk assessment",
    "3. Obtain verbal approval (document in CR) - document approver name, time, conditions",
    "4. Document expedited approval in CR - Approval type, Approvers, Approval time, Conditions"
]

for step in accelerated:
    ws12.cell(row=row, column=1, value=step)
    ws12.cell(row=row, column=1).border = thin_border
    row += 1

row += 1
add_header(ws12, "Compressed Deployment Timeline", row)
row += 1

timeline_headers = ["Phase", "Standard", "Emergency"]
for col, h in enumerate(timeline_headers, 1):
    ws12.cell(row=row, column=col, value=h)
    ws12.cell(row=row, column=col).font = header_font
    ws12.cell(row=row, column=col).fill = header_fill
    ws12.cell(row=row, column=col).border = thin_border
row += 1

timeline = [
    ("Approval", "5 days", "4-24 hours"),
    ("DEV Deployment", "1 day", "Same day"),
    ("PROD Deployment", "1-2 days", "4-12 hours"),
    ("Post-Deploy Validation", "Standard", "24-48 hours in TEST")
]

for t in timeline:
    for col, val in enumerate(t, 1):
        ws12.cell(row=row, column=col, value=val)
        ws12.cell(row=row, column=col).border = thin_border
    row += 1

row += 1
add_header(ws12, "Post-Incident Review Requirements", row)
row += 1

post_incident = [
    "1. Schedule Post-Incident Review (PIR) within 5 business days of deployment",
    "2. Document PIR agenda: Timeline, Effectiveness, Process improvements, Lessons learned",
    "3. Complete PIR documentation: Root cause, Response effectiveness, Deviations, Recommendations",
    "4. Submit PIR to CAB for review and archive for compliance"
]

for step in post_incident:
    ws12.cell(row=row, column=1, value=step)
    ws12.cell(row=row, column=1).border = thin_border
    row += 1

ws12.column_dimensions['A'].width = 25
ws12.column_dimensions['B'].width = 30
ws12.column_dimensions['C'].width = 30

# Save the workbook
wb.save('/home/aaron/template/Patch_Management_Complete.xlsx')
print("Excel file created successfully!")
